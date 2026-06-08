"""
Service for importing and exporting questions and exams via CSV and Excel formats.
Supports bulk operations for question banks and exam templates.
"""

import csv
import io
from datetime import datetime
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json

from models.question_model import _get_tags, create_question, get_question_by_id, list_questions

from models.exam_model import create_exam, get_exam_by_id


# Keep legacy compatibility with older service code expecting QuestionModel/ExamModel
from models import question_model as _question_model
from models import exam_model as _exam_model


QuestionModel = _question_model
ExamModel = _exam_model


from schemas.question_schema import QuestionCreate
from schemas.exam_schema import ExamCreate





class ImportExportService:

    """Handle import and export of questions and exams."""

    @staticmethod
    async def export_questions_csv(question_ids: List[int] = None) -> str:
        """
        Export questions to CSV format.
        
        Args:
            question_ids: Optional list of question IDs to export. If None, export all.
        
        Returns:
            CSV content as string.
        """
        question_model = QuestionModel()
        
        if question_ids:
            questions = []
            for qid in question_ids:
                q = await question_model.get_question(qid)
                if q:
                    questions.append(q)
        else:
            questions = await question_model.get_all_questions()
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            'id', 'title', 'description', 'type', 'difficulty', 'tags',
            'image_url', 'media_url', 'attachment_url', 'created_by'
        ])
        writer.writeheader()
        
        for q in questions:
            writer.writerow({
                'id': q.get('id', ''),
                'title': q.get('title', ''),
                'description': q.get('description', ''),
                'type': q.get('type', 'mcq'),
                'difficulty': q.get('difficulty', ''),
                'tags': json.dumps(q.get('tags', [])) if isinstance(q.get('tags'), list) else q.get('tags', ''),
                'image_url': q.get('image_url', ''),
                'media_url': q.get('media_url', ''),
                'attachment_url': q.get('attachment_url', ''),
                'created_by': q.get('created_by', '')
            })
        
        return output.getvalue()

    @staticmethod
    async def export_questions_excel(question_ids: List[int] = None) -> bytes:
        """
        Export questions to Excel format with formatting.
        
        Args:
            question_ids: Optional list of question IDs to export. If None, export all.
        
        Returns:
            Excel file content as bytes.
        """
        question_model = QuestionModel()
        
        if question_ids:
            questions = []
            for qid in question_ids:
                q = await question_model.get_question(qid)
                if q:
                    questions.append(q)
        else:
            questions = await question_model.get_all_questions()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Questions"
        
        # Header row
        headers = ['ID', 'Title', 'Description', 'Type', 'Difficulty', 'Tags', 
                   'Image URL', 'Media URL', 'Attachment URL', 'Created By']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data rows
        for q in questions:
            ws.append([
                q.get('id', ''),
                q.get('title', ''),
                q.get('description', ''),
                q.get('type', 'mcq'),
                q.get('difficulty', ''),
                json.dumps(q.get('tags', [])) if isinstance(q.get('tags'), list) else q.get('tags', ''),
                q.get('image_url', ''),
                q.get('media_url', ''),
                q.get('attachment_url', ''),
                q.get('created_by', '')
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 15
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    async def import_questions_csv(file_content: str, user_id: int) -> Tuple[int, List[str]]:
        """
        Import questions from CSV format.
        
        Args:
            file_content: CSV file content as string.
            user_id: ID of user creating the questions.
        
        Returns:
            Tuple of (count_imported, list_of_errors).
        """
        question_model = QuestionModel()
        reader = csv.DictReader(io.StringIO(file_content))
        count_imported = 0
        errors = []
        
        for row_num, row in enumerate(reader, start=2):  # Start at 2 (after header)
            try:
                # Parse tags if it's a JSON string
                tags = []
                if row.get('tags'):
                    try:
                        tags = json.loads(row['tags'])
                    except json.JSONDecodeError:
                        tags = [t.strip() for t in row['tags'].split(',')]
                
                question_data = QuestionCreate(
                    title=row.get('title', '').strip(),
                    description=row.get('description', '').strip(),
                    type=row.get('type', 'mcq').strip(),
                    difficulty=row.get('difficulty', 'medium').strip(),
                    tags=tags,
                    image_url=row.get('image_url', '').strip() or None,
                    media_url=row.get('media_url', '').strip() or None,
                    attachment_url=row.get('attachment_url', '').strip() or None,
                )
                
                # Create question
                await question_model.create_question(
                    title=question_data.title,
                    description=question_data.description,
                    question_type=question_data.type,
                    difficulty=question_data.difficulty,
                    tags=question_data.tags,
                    created_by=user_id,
                    image_url=question_data.image_url,
                    media_url=question_data.media_url,
                    attachment_url=question_data.attachment_url,
                )
                count_imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        return count_imported, errors

    @staticmethod
    async def import_questions_excel(file_bytes: bytes, user_id: int) -> Tuple[int, List[str]]:
        """
        Import questions from Excel format.
        
        Args:
            file_bytes: Excel file content as bytes.
            user_id: ID of user creating the questions.
        
        Returns:
            Tuple of (count_imported, list_of_errors).
        """
        question_model = QuestionModel()
        errors = []
        count_imported = 0
        
        try:
            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            
            # Skip header row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not row[0]:  # Skip empty rows
                        continue
                    
                    # Parse tags
                    tags = []
                    tags_value = row[5] if len(row) > 5 else None
                    if tags_value:
                        try:
                            tags = json.loads(tags_value)
                        except (json.JSONDecodeError, TypeError):
                            tags = [t.strip() for t in str(tags_value).split(',') if t.strip()]
                    
                    question_data = QuestionCreate(
                        title=str(row[1] or '').strip(),
                        description=str(row[2] or '').strip(),
                        type=str(row[3] or 'mcq').strip(),
                        difficulty=str(row[4] or 'medium').strip(),
                        tags=tags,
                        image_url=str(row[6] or '').strip() or None if len(row) > 6 else None,
                        media_url=str(row[7] or '').strip() or None if len(row) > 7 else None,
                        attachment_url=str(row[8] or '').strip() or None if len(row) > 8 else None,
                    )
                    
                    # Create question
                    await question_model.create_question(
                        title=question_data.title,
                        description=question_data.description,
                        question_type=question_data.type,
                        difficulty=question_data.difficulty,
                        tags=question_data.tags,
                        created_by=user_id,
                        image_url=question_data.image_url,
                        media_url=question_data.media_url,
                        attachment_url=question_data.attachment_url,
                    )
                    count_imported += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
        
        except Exception as e:
            errors.append(f"Excel parsing error: {str(e)}")
        
        return count_imported, errors

    @staticmethod
    async def export_exams_csv(exam_ids: List[int] = None) -> str:
        """
        Export exams to CSV format.
        
        Args:
            exam_ids: Optional list of exam IDs to export. If None, export all.
        
        Returns:
            CSV content as string.
        """
        exam_model = ExamModel()
        
        if exam_ids:
            exams = []
            for eid in exam_ids:
                e = await exam_model.get_exam(eid)
                if e:
                    exams.append(e)
        else:
            exams = await exam_model.get_all_exams()
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            'id', 'title', 'description', 'duration_minutes', 'total_marks',
            'passing_marks', 'secure_mode', 'is_public', 'created_by'
        ])
        writer.writeheader()
        
        for e in exams:
            writer.writerow({
                'id': e.get('id', ''),
                'title': e.get('title', ''),
                'description': e.get('description', ''),
                'duration_minutes': e.get('duration_minutes', ''),
                'total_marks': e.get('total_marks', ''),
                'passing_marks': e.get('passing_marks', ''),
                'secure_mode': e.get('secure_mode', 0),
                'is_public': e.get('is_public', 0),
                'created_by': e.get('created_by', '')
            })
        
        return output.getvalue()

    @staticmethod
    async def export_exams_excel(exam_ids: List[int] = None) -> bytes:
        """
        Export exams to Excel format with formatting.
        
        Args:
            exam_ids: Optional list of exam IDs to export. If None, export all.
        
        Returns:
            Excel file content as bytes.
        """
        exam_model = ExamModel()
        
        if exam_ids:
            exams = []
            for eid in exam_ids:
                e = await exam_model.get_exam(eid)
                if e:
                    exams.append(e)
        else:
            exams = await exam_model.get_all_exams()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Exams"
        
        # Header row
        headers = ['ID', 'Title', 'Description', 'Duration (mins)', 'Total Marks', 
                   'Passing Marks', 'Secure Mode', 'Is Public', 'Created By']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data rows
        for e in exams:
            ws.append([
                e.get('id', ''),
                e.get('title', ''),
                e.get('description', ''),
                e.get('duration_minutes', ''),
                e.get('total_marks', ''),
                e.get('passing_marks', ''),
                'Yes' if e.get('secure_mode') else 'No',
                'Yes' if e.get('is_public') else 'No',
                e.get('created_by', '')
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    async def import_exams_csv(file_content: str, user_id: int) -> Tuple[int, List[str]]:
        """
        Import exams from CSV format.
        
        Args:
            file_content: CSV file content as string.
            user_id: ID of user creating the exams.
        
        Returns:
            Tuple of (count_imported, list_of_errors).
        """
        exam_model = ExamModel()
        reader = csv.DictReader(io.StringIO(file_content))
        count_imported = 0
        errors = []
        
        for row_num, row in enumerate(reader, start=2):
            try:
                exam_data = ExamCreate(
                    title=row.get('title', '').strip(),
                    description=row.get('description', '').strip(),
                    duration_minutes=int(row.get('duration_minutes', 60)),
                    total_marks=int(row.get('total_marks', 100)),
                    passing_marks=int(row.get('passing_marks', 50)),
                    secure_mode=row.get('secure_mode', '0').lower() in ('1', 'true', 'yes'),
                    is_public=row.get('is_public', '1').lower() in ('1', 'true', 'yes'),
                )
                
                # Create exam
                await exam_model.create_exam(
                    title=exam_data.title,
                    description=exam_data.description,
                    duration_minutes=exam_data.duration_minutes,
                    total_marks=exam_data.total_marks,
                    passing_marks=exam_data.passing_marks,
                    created_by=user_id,
                    secure_mode=exam_data.secure_mode,
                    is_public=exam_data.is_public,
                )
                count_imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        return count_imported, errors

    @staticmethod
    async def import_exams_excel(file_bytes: bytes, user_id: int) -> Tuple[int, List[str]]:
        """
        Import exams from Excel format.
        
        Args:
            file_bytes: Excel file content as bytes.
            user_id: ID of user creating the exams.
        
        Returns:
            Tuple of (count_imported, list_of_errors).
        """
        exam_model = ExamModel()
        errors = []
        count_imported = 0
        
        try:
            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            
            # Skip header row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not row[0]:  # Skip empty rows
                        continue
                    
                    exam_data = ExamCreate(
                        title=str(row[1] or '').strip(),
                        description=str(row[2] or '').strip(),
                        duration_minutes=int(row[3] or 60),
                        total_marks=int(row[4] or 100),
                        passing_marks=int(row[5] or 50),
                        secure_mode=str(row[6] or 'No').lower() in ('yes', '1', 'true'),
                        is_public=str(row[7] or 'Yes').lower() in ('yes', '1', 'true'),
                    )
                    
                    # Create exam
                    await exam_model.create_exam(
                        title=exam_data.title,
                        description=exam_data.description,
                        duration_minutes=exam_data.duration_minutes,
                        total_marks=exam_data.total_marks,
                        passing_marks=exam_data.passing_marks,
                        created_by=user_id,
                        secure_mode=exam_data.secure_mode,
                        is_public=exam_data.is_public,
                    )
                    count_imported += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
        
        except Exception as e:
            errors.append(f"Excel parsing error: {str(e)}")
        
        return count_imported, errors
